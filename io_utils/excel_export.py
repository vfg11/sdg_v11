"""
io_utils/excel_export.py — Exportar señales de todas las hojas.
SlotData ya no tiene campo signal (eliminado en v5).
"""
from __future__ import annotations
from pathlib import Path
from model import DocumentData

HEADERS = ['#', 'KKS Entrada', 'Ref. Origen',
           'KKS Salida',  'Ref. Destino', 'Enlazado']


def export_excel(document: DocumentData, path: str | Path):
    path = Path(path)
    try:
        import openpyxl
        _to_xlsx(document, path.with_suffix('.xlsx'))
    except ImportError:
        _to_csv(document, path.with_suffix('.csv'))


def _to_xlsx(document: DocumentData, path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb  = openpyxl.Workbook()
    wb.remove(wb.active)
    thin = Side(border_style='thin', color='AABBCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_idx, (s, _g) in enumerate(document.flat_sheets()):
        num = document.sheet_ref(sheet_idx)
        ws  = wb.create_sheet(title=f'{num}_{s.sheet_name[:18]}')
        tb  = document.title_block

        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = (f"{tb.title}  |  {s.sheet_name}  "
                    f"|  Doc: {tb.doc_number}  Rev: {tb.revision}")
        ws['A1'].font      = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        ws['A1'].fill      = PatternFill('solid', fgColor='1A2A4A')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Cabecera
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font      = Font(name='Calibri', bold=True, size=9)
            c.fill      = PatternFill('solid', fgColor='D0DAF0')
            c.border    = brd
            c.alignment = Alignment(horizontal='center')

        # Datos por cajón
        for i in range(s.num_slots):
            sl = s.slots_left[i]  if i < len(s.slots_left)  else None
            sr = s.slots_right[i] if i < len(s.slots_right) else None
            linked = ''
            if sl and sl.linked_sheet >= 0:
                linked = f'E→{sl.sub_text}'
            elif sr and sr.linked_sheet >= 0:
                linked = f'S→{sr.sub_text}'
            row_data = [
                i + 1,
                sl.kks      if sl else '',
                sl.sub_text if sl else '',
                sr.kks      if sr else '',
                sr.sub_text if sr else '',
                linked,
            ]
            fill = PatternFill('solid', fgColor='F5F8FF' if i % 2 == 0 else 'FFFFFF')
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=i + 3, column=col, value=val)
                c.font   = Font(name='Courier New', size=8)
                c.border = brd
                c.fill   = fill

        for col_letter, w in zip('ABCDEF', [4, 20, 12, 20, 12, 12]):
            ws.column_dimensions[col_letter].width = w

    # Hoja de bloques
    ws_blk = wb.create_sheet('Bloques')
    ws_blk.append(['Hoja', 'Nº', 'ID', 'Tipo', 'KKS', 'Etiqueta', 'Entradas', 'Salidas'])
    for si, (s, _g) in enumerate(document.flat_sheets()):
        num = document.sheet_ref(si)
        for b in s.blocks:
            ws_blk.append([s.sheet_name, num, b.block_id[:8], b.type_id,
                            b.kks, b.label or b.display_name(),
                            len(b.inputs), len(b.outputs)])

    wb.save(path)


def _to_csv(document: DocumentData, path: Path):
    import csv
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        for si, (s, _g) in enumerate(document.flat_sheets()):
            num = document.sheet_ref(si)
            w.writerow([f'=== Hoja {num}: {s.sheet_name} ==='])
            w.writerow(HEADERS)
            for i in range(s.num_slots):
                sl = s.slots_left[i]  if i < len(s.slots_left)  else None
                sr = s.slots_right[i] if i < len(s.slots_right) else None
                linked = ''
                if sl and sl.linked_sheet >= 0: linked = f'E→{sl.sub_text}'
                elif sr and sr.linked_sheet >= 0: linked = f'S→{sr.sub_text}'
                w.writerow([i+1,
                    sl.kks if sl else '', sl.sub_text if sl else '',
                    sr.kks if sr else '', sr.sub_text if sr else '',
                    linked])
            w.writerow([])
