"""
io_utils/bulk_clone.py — Clonación masiva de grupos desde fichero Excel.

Estructura del fichero:
  - Filas 1-9: ignoradas (zona libre para cabeceras / instrucciones)
  - A partir de fila 10: pares de filas
      Fila impar (origen):  col1=hoja, col2=kks, col3=descripción, col4=sistema,
                            col5..=patrones de búsqueda
      Fila par  (destino):  col1=hoja_destino, col2=kks_nuevo, col3=desc_nueva,
                            col4=sistema_nuevo, col5..=cadenas de reemplazo

  Los pares (búsqueda, reemplazo) se forman por posición de columna:
  col5_origen↔col5_destino, col6_origen↔col6_destino, ...
  Columnas de búsqueda vacías se omiten.

API pública:
  parse_bulk_excel(path)  → list[BulkOp]  (o lanza BulkParseError)
  validate_bulk(ops, doc) → list[str]     (errores de validación)
  execute_bulk(ops, doc_id, on_progress)  → str nuevo último group_id
                                            (dentro de una transacción; quien
                                             llama hace commit o rollback)
"""
from __future__ import annotations
import uuid
from dataclasses import dataclass, field

FIRST_DATA_ROW = 10   # filas 1-9 descartadas


class BulkParseError(Exception):
    pass


@dataclass
class BulkOp:
    """Una operación de clonado: origen identificado por kks, destino con override."""
    row_pair:    int              # número de fila del par (conteo desde 1)
    src_kks:     str              # KKS del grupo origen
    dst_sheet:   int              # número de hoja destino (override_base)
    dst_kks:     str
    dst_desc:    str
    dst_system:  str
    rules:       list[tuple[str, str]] = field(default_factory=list)
    # Resuelto durante validación:
    src_group_id: str = ''


# ── Parser ───────────────────────────────────────────────────────────────────

def parse_bulk_excel(path: str) -> list[BulkOp]:
    """
    Lee el fichero Excel y devuelve la lista de BulkOp.
    Lanza BulkParseError con mensaje descriptivo si algo está mal.
    """
    try:
        import openpyxl
    except ImportError:
        raise BulkParseError("openpyxl no está instalado. "
                             "Instálalo con: pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise BulkParseError(f"No se pudo abrir el fichero: {e}")

    ws = wb.active
    # Leer todas las filas como listas de valores (None si vacía)
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    wb.close()

    # Filas de datos: desde índice FIRST_DATA_ROW-1 (0-based)
    data_rows = rows[FIRST_DATA_ROW - 1:]

    # Eliminar filas completamente vacías al final
    while data_rows and all(v is None or str(v).strip() == '' for v in data_rows[-1]):
        data_rows.pop()

    if not data_rows:
        raise BulkParseError(
            f"El fichero no contiene datos a partir de la fila {FIRST_DATA_ROW}.")

    if len(data_rows) % 2 != 0:
        raise BulkParseError(
            f"El número de filas de datos (a partir de fila {FIRST_DATA_ROW}) "
            f"debe ser par (pares origen/destino). "
            f"Se encontraron {len(data_rows)} filas.")

    def _cell(row, col):
        """col es 0-based. Devuelve str limpio o ''."""
        if col >= len(row):
            return ''
        v = row[col]
        return str(v).strip() if v is not None else ''

    ops: list[BulkOp] = []
    pair_num = 0
    for i in range(0, len(data_rows), 2):
        pair_num += 1
        abs_row = FIRST_DATA_ROW + i  # número de fila en el Excel (1-based)
        src_row = data_rows[i]
        dst_row = data_rows[i + 1]

        # Extraer campos fijos
        src_kks  = _cell(src_row, 1)   # col2 (0-based=1)
        dst_sheet_raw = _cell(dst_row, 0)
        dst_kks  = _cell(dst_row, 1)
        dst_desc = _cell(dst_row, 2)
        dst_sys  = _cell(dst_row, 3)

        # Validaciones básicas de formato
        if not src_kks:
            raise BulkParseError(
                f"Fila {abs_row} (origen del par {pair_num}): "
                f"KKS vacío.")

        if not dst_sheet_raw:
            raise BulkParseError(
                f"Fila {abs_row+1} (destino del par {pair_num}): "
                f"número de hoja vacío.")

        try:
            dst_sheet = int(float(dst_sheet_raw))
            if dst_sheet < 1:
                raise ValueError
        except ValueError:
            raise BulkParseError(
                f"Fila {abs_row+1} (destino del par {pair_num}): "
                f"número de hoja '{dst_sheet_raw}' no es un entero positivo.")

        if not dst_kks:
            raise BulkParseError(
                f"Fila {abs_row+1} (destino del par {pair_num}): "
                f"KKS destino vacío.")

        # Reglas: columnas 5+ (0-based=4+)
        rules: list[tuple[str, str]] = []
        max_col = max(len(src_row), len(dst_row))
        for col in range(4, max_col):
            pat = _cell(src_row, col)
            rep = _cell(dst_row, col)
            if pat:   # solo si hay patrón de búsqueda
                rules.append((pat, rep))

        ops.append(BulkOp(
            row_pair=pair_num,
            src_kks=src_kks,
            dst_sheet=dst_sheet,
            dst_kks=dst_kks,
            dst_desc=dst_desc,
            dst_system=dst_sys,
            rules=rules,
        ))

    return ops


# ── Copia temporal desde BD externa ──────────────────────────────────────────

_TMP_DOC = '__tmp_import__'

def _copy_group_to_mem(src_con, src_group_id: str) -> str:
    """
    Copia un grupo de src_con a _mem bajo doc_id=_TMP_DOC.
    Devuelve el nuevo group_id temporal.
    Tablas copiadas: groups, sheets, slots, blocks, block_ports,
                     connections, waypoints, branch_nodes, symbols,
                     notes, textboxes, slot_links.
    """
    from io_utils.db_io import get_mem
    mem = get_mem()

    # Eliminar restos de copias anteriores
    mem.execute("DELETE FROM groups WHERE doc_id=?", (_TMP_DOC,))
    mem.commit()

    tmp_gid = str(uuid.uuid4())

    # ── group ──────────────────────────────────────────────────────────
    g = src_con.execute(
        "SELECT system,description,kks,revision,date,sheet_number_base,order_idx "
        "FROM groups WHERE group_id=?", (src_group_id,)).fetchone()
    if not g:
        raise ValueError(f"Grupo '{src_group_id}' no encontrado en la BD origen.")
    mem.execute(
        "INSERT INTO groups (group_id,doc_id,order_idx,system,description,"
        "kks,revision,date,sheet_number_base) VALUES (?,?,?,?,?,?,?,?,?)",
        (tmp_gid, _TMP_DOC, g[6], g[0], g[1], g[2], g[3], g[4], g[5]))

    # ── sheets + contenido ─────────────────────────────────────────────
    sheets = src_con.execute(
        "SELECT sheet_id,order_idx,num_slots,sheet_name,sheet_title,sheet_number "
        "FROM sheets WHERE group_id=? ORDER BY order_idx",
        (src_group_id,)).fetchall()

    sheet_map: dict[str, str] = {}   # old_sid → new_sid
    slot_map:  dict[str, str] = {}   # old_slot_id → new_slot_id

    for sh in sheets:
        old_sid, oidx, ns, sname, stitle, snum = sh
        new_sid = str(uuid.uuid4())
        sheet_map[old_sid] = new_sid
        mem.execute(
            "INSERT INTO sheets (sheet_id,group_id,order_idx,num_slots,"
            "sheet_name,sheet_title,sheet_number) VALUES (?,?,?,?,?,?,?)",
            (new_sid, tmp_gid, oidx, ns, sname, stitle, snum))

        # slots
        slots = src_con.execute(
            "SELECT slot_id,side,slot_index,description,signal_desc,"
            "kks,kks2,sub_text,signal_type FROM slots WHERE sheet_id=?",
            (old_sid,)).fetchall()
        for sl in slots:
            new_slid = str(uuid.uuid4())
            slot_map[sl[0]] = new_slid
            mem.execute(
                "INSERT INTO slots (slot_id,sheet_id,side,slot_index,"
                "description,signal_desc,kks,kks2,sub_text,signal_type) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (new_slid, new_sid, sl[1], sl[2], sl[3], sl[4],
                 sl[5], sl[6], sl[7], sl[8]))

        # blocks + ports
        block_map: dict[str, str] = {}
        port_map:  dict[str, str] = {}
        blocks = src_con.execute(
            "SELECT block_id,block_type,x,y,inscription,port_side,"
            "negated,signal_type,kks,num_inputs FROM blocks WHERE sheet_id=?",
            (old_sid,)).fetchall()
        for bl in blocks:
            new_bid = str(uuid.uuid4())
            block_map[bl[0]] = new_bid
            mem.execute(
                "INSERT INTO blocks (block_id,sheet_id,block_type,x,y,"
                "inscription,port_side,negated,signal_type,kks,num_inputs) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (new_bid, new_sid, bl[1], bl[2], bl[3], bl[4],
                 bl[5], bl[6], bl[7], bl[8], bl[9]))
            ports = src_con.execute(
                "SELECT port_id,side,label FROM block_ports WHERE block_id=?",
                (bl[0],)).fetchall()
            for pt in ports:
                new_pid = str(uuid.uuid4())
                port_map[pt[0]] = new_pid
                mem.execute(
                    "INSERT INTO block_ports (port_id,block_id,side,label) "
                    "VALUES (?,?,?,?)",
                    (new_pid, new_bid, pt[1], pt[2]))

        # branch_nodes primero — sus IDs se usan como src/dst en connections
        branch_map: dict[str, str] = {}
        try:
            bns = src_con.execute(
                "SELECT branch_id,parent_conn_id,x,y FROM branch_nodes "
                "WHERE sheet_id=?", (old_sid,)).fetchall()
        except Exception:
            bns = []
        pending_bns = list(bns)
        for bn in pending_bns:
            branch_map[bn[0]] = str(uuid.uuid4())

        # connections
        conn_map: dict[str, str] = {}
        try:
            conns = src_con.execute(
                "SELECT conn_id,src,src_is_slot,dst,dst_is_slot,"
                "signal_type FROM connections WHERE sheet_id=?",
                (old_sid,)).fetchall()
        except Exception:
            conns = []
        for cn in conns:
            new_cid = str(uuid.uuid4())
            conn_map[cn[0]] = new_cid
            def _remap(ref, is_slot,
                       _sm=slot_map, _bm=block_map, _pm=port_map, _brm=branch_map):
                if is_slot: return _sm.get(ref, ref)
                return _bm.get(ref) or _pm.get(ref) or _brm.get(ref) or ref
            mem.execute(
                "INSERT INTO connections (conn_id,sheet_id,src,src_is_slot,"
                "dst,dst_is_slot,signal_type) VALUES (?,?,?,?,?,?,?)",
                (new_cid, new_sid,
                 _remap(cn[1], bool(cn[2])), cn[2],
                 _remap(cn[3], bool(cn[4])), cn[4], cn[5]))
            # waypoints
            try:
                wps = src_con.execute(
                    "SELECT x,y,ordinal FROM waypoints WHERE conn_id=?",
                    (cn[0],)).fetchall()
                for wp in wps:
                    mem.execute(
                        "INSERT INTO waypoints (wp_id,conn_id,x,y,ordinal) "
                        "VALUES (?,?,?,?,?)",
                        (str(uuid.uuid4()), new_cid, wp[0], wp[1], wp[2]))
            except Exception:
                pass

        # Insertar branch_nodes con parent_conn ya remapeado
        for bn in pending_bns:
            new_pc = conn_map.get(bn[1], bn[1])
            mem.execute(
                "INSERT INTO branch_nodes "
                "(branch_id,sheet_id,parent_conn_id,x,y) VALUES (?,?,?,?,?)",
                (branch_map[bn[0]], new_sid, new_pc, bn[2], bn[3]))

        # symbols
        try:
            syms = src_con.execute(
                "SELECT sym_type,port_side,x,y,inscription,sym_id "
                "FROM symbols WHERE sheet_id=?", (old_sid,)).fetchall()
            for sy in syms:
                mem.execute(
                    "INSERT INTO symbols "
                    "(sym_id,sheet_id,sym_type,port_side,x,y,inscription) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (str(uuid.uuid4()), new_sid,
                     sy[0], sy[1], sy[2], sy[3], sy[4]))
        except Exception:
            pass

        # notes
        try:
            notes = src_con.execute(
                "SELECT text,x,y,font_size_px,text_width "
                "FROM notes WHERE sheet_id=?", (old_sid,)).fetchall()
            for nt in notes:
                mem.execute(
                    "INSERT INTO notes "
                    "(note_id,sheet_id,text,x,y,font_size_px,text_width) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (str(uuid.uuid4()), new_sid,
                     nt[0], nt[1], nt[2], nt[3], nt[4]))
        except Exception:
            pass

        # textboxes
        try:
            tbs = src_con.execute(
                "SELECT text,x,y,w,h,signal_type,port_side "
                "FROM textboxes WHERE sheet_id=?", (old_sid,)).fetchall()
            for tb in tbs:
                mem.execute(
                    "INSERT INTO textboxes "
                    "(box_id,sheet_id,text,x,y,w,h,signal_type,port_side) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (str(uuid.uuid4()), new_sid,
                     tb[0], tb[1], tb[2], tb[3], tb[4], tb[5], tb[6]))
        except Exception:
            pass

    # slot_links (se refieren a slot_ids de todo el grupo)
    old_sids_tuple = tuple(sheet_map.keys())
    if old_sids_tuple:
        ph = ','.join('?' * len(old_sids_tuple))
        try:
            links = src_con.execute(
                f"SELECT src_slot_id,dst_slot_id FROM slot_links "
                f"WHERE src_slot_id IN ({ph}) OR dst_slot_id IN ({ph})",
                old_sids_tuple + old_sids_tuple).fetchall()
            for lk in links:
                ns = slot_map.get(lk[0])
                nd = slot_map.get(lk[1])
                if ns and nd:
                    mem.execute(
                        "INSERT INTO slot_links (link_id,src_slot_id,dst_slot_id) "
                        "VALUES (?,?,?)",
                        (str(uuid.uuid4()), ns, nd))
        except Exception:
            pass

    mem.commit()
    return tmp_gid


def _cleanup_tmp(mem_con):
    """Elimina el grupo temporal creado por _copy_group_to_mem."""
    mem_con.execute("DELETE FROM groups WHERE doc_id=?", (_TMP_DOC,))
    mem_con.commit()


# ── Validación ────────────────────────────────────────────────────────────────

def validate_bulk(ops: list[BulkOp], doc, src_con=None) -> list[str]:
    """
    Valida la lista de operaciones contra el documento.
    Si src_con es una conexión SQLite externa, los KKS origen se buscan
    en esa BD en lugar del documento actual.
    Resuelve src_group_id en cada BulkOp.
    Devuelve lista de mensajes de error (vacía = todo OK).
    """
    from io_utils.clone_group import validate_rules

    errors: list[str] = []

    # Índice KKS → group_id
    kks_to_gid: dict[str, str] = {}

    if src_con is not None:
        # Buscar KKS en la BD externa (un solo documento por fichero .sde)
        rows = src_con.execute(
            "SELECT group_id, kks FROM groups").fetchall()
        for gid, kks in rows:
            k = (kks or '').strip()
            if k:
                if k in kks_to_gid:
                    errors.append(
                        f"La BD origen tiene dos grupos con KKS '{k}'. "
                        f"La importación masiva requiere KKS únicos.")
                else:
                    kks_to_gid[k] = gid
    else:
        for g in doc.groups:
            k = (g.kks or '').strip()
            if k:
                if k in kks_to_gid:
                    errors.append(
                        f"El documento tiene dos grupos con KKS '{k}'. "
                        f"La importación masiva requiere KKS únicos.")
                else:
                    kks_to_gid[k] = g.group_id

    for op in ops:
        gid = kks_to_gid.get(op.src_kks, '')
        if not gid:
            src_label = "la BD origen" if src_con else "el documento"
            errors.append(
                f"Par {op.row_pair}: KKS origen '{op.src_kks}' "
                f"no existe en {src_label}.")
        else:
            op.src_group_id = gid

        rule_errors = validate_rules(op.rules)
        for re_msg in rule_errors:
            errors.append(f"Par {op.row_pair}: regla inválida — {re_msg}")

    return errors


# ── Ejecución ─────────────────────────────────────────────────────────────────

def execute_bulk(ops: list[BulkOp], doc_id: str,
                 on_progress=None, src_con=None) -> list[str]:
    """
    Ejecuta las clonaciones en secuencia DENTRO DE LA CONEXIÓN ACTIVA.
    No hace commit ni rollback — responsabilidad del llamador.

    on_progress(current, total, msg) — callback opcional para la UI.
    Devuelve lista de nuevos group_ids creados.
    """
    from io_utils.clone_group import clone_group_sql, apply_kks_autolink
    from io_utils.db_io import get_mem

    con = get_mem()
    new_gids: list[str] = []
    total = len(ops)

    for i, op in enumerate(ops):
        if on_progress:
            on_progress(i, total,
                        f"Clonando par {op.row_pair}: '{op.src_kks}' → "
                        f"'{op.dst_kks}' (H{op.dst_sheet:02d})")

        # Si hay BD origen externa, copiar el grupo temporalmente a _mem
        effective_src_gid = op.src_group_id
        if src_con is not None:
            effective_src_gid = _copy_group_to_mem(src_con, op.src_group_id)

        # Construir reglas: además de las del Excel, sustituir kks/desc/sistema
        # del grupo con los valores destino
        src_row = con.execute(
            "SELECT kks, description, system FROM groups WHERE group_id=?",
            (effective_src_gid,)).fetchone()
        auto_rules: list[tuple[str, str]] = []
        if src_row:
            if src_row[0] and op.dst_kks:
                auto_rules.append((src_row[0], op.dst_kks))
            if src_row[1] and op.dst_desc:
                auto_rules.append((src_row[1], op.dst_desc))
            if src_row[2] and op.dst_system:
                auto_rules.append((src_row[2], op.dst_system))

        # Reglas del Excel van DESPUÉS de las automáticas (más específicas primero)
        all_rules = auto_rules + op.rules

        # Insertar al final y luego clone_group_sql reposiciona con override_base
        new_gid = clone_group_sql(
            src_group_id=effective_src_gid,
            insert_before_group_id='__END__',
            doc_id=doc_id,
            rules=all_rules,
            override_base=op.dst_sheet,
        )
        new_gids.append(new_gid)

        # Limpiar copia temporal si la había
        if src_con is not None:
            _cleanup_tmp(con)

    if on_progress:
        on_progress(total, total, "Completado.")

    return new_gids


# ── Exportación de plantilla ─────────────────────────────────────────────────

def export_template_excel(doc, group_ids: list[str], path: str):
    """
    Genera un fichero Excel de plantilla con los grupos seleccionados.
    Filas 1-9: instrucciones.
    A partir de fila 10: una fila por grupo con hoja, kks, descripción, sistema.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl no está instalado.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla clonación"

    # ── Estilos ──────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="2E4057")
    inst_fill = PatternFill("solid", fgColor="E8F4FD")
    inst_font = Font(color="1A3A5C", size=10, italic=True)
    src_fill  = PatternFill("solid", fgColor="EBF5EB")   # verde claro = origen
    dst_fill  = PatternFill("solid", fgColor="FFF9E6")   # amarillo claro = destino
    bold      = Font(bold=True)

    def _set(row, col, value, font=None, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font  = font
        if fill:  c.fill  = fill
        if align: c.alignment = align
        return c

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ── Filas 1-9: instrucciones ──────────────────────────────────────────
    ws.merge_cells('A1:H1')
    _set(1, 1, "PLANTILLA DE CLONACIÓN MASIVA DE GRUPOS",
         font=Font(bold=True, color="FFFFFF", size=14),
         fill=PatternFill("solid", fgColor="1A3A5C"),
         align=center)
    ws.row_dimensions[1].height = 28

    instrucciones = [
        (2, "INSTRUCCIONES DE USO"),
        (3, "• A partir de la fila 10 se definen pares de clonación (fila origen + fila destino)."),
        (4, "• Fila ORIGEN (fondo verde): col1=hoja actual, col2=KKS del grupo a clonar, "
             "col3=descripción, col4=sistema. "
             "Columnas 5 en adelante: patrones de búsqueda (con * y ? como comodines)."),
        (5, "• Fila DESTINO (fondo amarillo): col1=hoja donde insertar el clon, "
             "col2=nuevo KKS, col3=nueva descripción, col4=nuevo sistema. "
             "Columnas 5 en adelante: cadenas de reemplazo (misma posición que la fila origen)."),
        (6, "• Las sustituciones de KKS, descripción y sistema se aplican automáticamente. "
             "Las columnas 5+ permiten sustituciones adicionales."),
        (7, "• Los pares se procesan en orden. Si dos pares piden la misma hoja destino, "
             "el primero se desplaza al insertar el segundo."),
        (8, "• Filas 1-9 son ignoradas al importar. No modifiques la estructura de columnas."),
        (9, None),
    ]
    for row, txt in instrucciones:
        if txt:
            ws.merge_cells(f'A{row}:H{row}')
            _set(row, 1, txt,
                 font=Font(bold=(row == 2), color="1A3A5C" if row > 2 else "1A3A5C",
                           size=10, italic=(row > 2)),
                 fill=inst_fill if row > 2 else PatternFill("solid", fgColor="C5DCF0"),
                 align=left)
        ws.row_dimensions[row].height = 22 if row > 2 else 20

    # ── Cabecera de columnas (dentro de fila 9 mezclamos una fila extra) ──
    # En realidad ponemos la cabecera en fila 9
    headers = ["Hoja", "KKS", "Descripción", "Sistema",
               "Búsqueda/Reemplazo 1", "Búsqueda/Reemplazo 2",
               "Búsqueda/Reemplazo 3", "Búsqueda/Reemplazo 4"]
    for ci, h in enumerate(headers, 1):
        _set(9, ci, h, font=hdr_font,
             fill=PatternFill("solid", fgColor="4A6FA5"),
             align=center)
    ws.row_dimensions[9].height = 20

    # Anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    for col in ['E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 22

    # ── Datos de grupos seleccionados (a partir de fila 10) ───────────────
    # Construir índice group_id → group
    gid_set = set(group_ids)
    selected = [g for g in doc.groups if g.group_id in gid_set]

    excel_row = FIRST_DATA_ROW
    for g in selected:
        base = g.sheet_number_base
        # Fila origen (fondo verde)
        vals_src = [base, g.kks or '', g.description or '', g.system or '']
        for ci, v in enumerate(vals_src, 1):
            _set(excel_row, ci, v, fill=src_fill,
                 font=Font(bold=(ci == 2)),
                 align=center if ci == 1 else left)
        # Celdas de búsqueda vacías (col 5-8) con fondo verde y placeholder
        for ci in range(5, 9):
            c = ws.cell(row=excel_row, column=ci, value='')
            c.fill = src_fill
            c.alignment = left
        ws.row_dimensions[excel_row].height = 18
        excel_row += 1

        # Fila destino (fondo amarillo) — el usuario rellena
        _set(excel_row, 1, base, fill=dst_fill, align=center)
        _set(excel_row, 2, g.kks or '',         fill=dst_fill, align=left,
             font=Font(bold=True))
        _set(excel_row, 3, g.description or '', fill=dst_fill, align=left)
        _set(excel_row, 4, g.system or '',      fill=dst_fill, align=left)
        for ci in range(5, 9):
            c = ws.cell(row=excel_row, column=ci, value='')
            c.fill = dst_fill
            c.alignment = left
        ws.row_dimensions[excel_row].height = 18
        excel_row += 1

    # Congelar filas 1-9
    ws.freeze_panes = 'A10'

    wb.save(path)
